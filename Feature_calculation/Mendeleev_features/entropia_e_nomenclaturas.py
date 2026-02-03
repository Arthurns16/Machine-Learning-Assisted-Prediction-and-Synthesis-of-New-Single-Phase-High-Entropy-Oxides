from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import math
from thermo_ml import parse

from importlib import reload


wb = load_workbook('Input_HEOs.xlsx')
ws = wb['Plan1']


contador = 0
linha_inicial = 4
linha_final = 151
numero_de_linhas = linha_final -linha_inicial




def atomos_efracoes_erudita (composto_bruto):
  

    reload(parse)
    CP = parse.ChemParser()
    tudo = CP.atoms(composto_bruto)
    tudo = str (tudo)

    vetor_atomosL = []
    vetor_fracoesL = []



    while tudo.find("'") > -1:

        indice_corte = tudo.index("'")
        tudo = tudo[indice_corte +1:]
        
        indice_fim_atomo = tudo.index("'")
        atomo = tudo[0:indice_fim_atomo]
        
        indice_inicio_fracao = tudo.index(": ") +2
        
        if tudo.find(",") >-1 :
            indice_fim_fracao = tudo.index(",") 
        else:
            indice_fim_fracao = tudo.index("}")
            
        fracao = tudo[indice_inicio_fracao:indice_fim_fracao] 
        
        tudo = tudo[indice_fim_fracao:]
        vetor_atomosL.append(atomo)
        vetor_fracoesL.append(fracao)
       
    
    atomos = vetor_atomosL.copy()
    fracoes = vetor_fracoesL.copy()
    
    del vetor_atomosL
    del vetor_fracoesL
    
    del CP
    del tudo
    

    return atomos , fracoes



def preencher_atomos_e_fracoes (linha_inicial,linha_final) :
   
   while linha_inicial < linha_final + 1:

        
        composto_bruto = ws["A"+ str(linha_inicial)].value
        composto_bruto =str(composto_bruto)
        

        ws["B"+ str(linha_inicial)] = str(atomos_efracoes_erudita(composto_bruto)[0])
        ws["C"+ str(linha_inicial)] = str(atomos_efracoes_erudita(composto_bruto)[1])

        
        
        linha_inicial =linha_inicial +1
        
        
       

  

def preencher_entropia_configuracional (linha_inicial,linha_final):

        
    while linha_inicial < linha_final + 1:

        
        composto_bruto = ws["A"+ str(linha_inicial)].value
        composto_bruto =str(composto_bruto)
       
        
        #OXIGENIO unico anion
    
    
        atomos = atomos_efracoes_erudita (composto_bruto)[0]
    
        n = len(atomos) - 1
    
        entropia_configuracional =  8.31*math.log(n) 
        
        ws["C"+ str(linha_inicial)] = entropia_configuracional
            
        
        linha_inicial = linha_inicial+1
      
        
    
    
    








preencher_entropia_configuracional(2,146) 
wb.save("Input_HEOs_TESTE.xlsx")







