from mendeleev import element
import numpy as np 
import math
import itertools
from itertools import permutations
from itertools import combinations 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import math
from thermo_ml import parse
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np 

def minimo (vetor) :
    return min(vetor)

def maximo (vetor) :
    return max(vetor)

def soma (vetor) :
    return sum(vetor)

def media (vetor) :
    return sum(vetor)/len(vetor)

def desvio (vetor) :
    return np.std(vetor)




def electronegativity_ghosh (elemento):

    L  =  element(elemento)
    electronegativity_ghosh = L.en_ghosh


    return electronegativity_ghosh
    


def vetor_electronegativity_ghosh_atomos (linha) :

   
    wb = load_workbook('Input_HEOs_TESTE.xlsx')
    ws = wb['Sheet1']
        
    atomos = ws["B"+ str(linha)].value
        

    atomos = atomos.replace (" ","")
    atomos = atomos.replace ("'","")
    atomos = atomos.replace ("[","")
    atomos = atomos.replace ("]","")
    atomos = atomos.split(",")

    vetor_propriedade_da_linha = []

    for elemento in atomos:
        try: 
            vetor_propriedade_da_linha.append (float(electronegativity_ghosh(elemento)))
        except:
            valor = 0
        
    return vetor_propriedade_da_linha




def criar_culunas_excel (propriedade) : 

    file_name = "Input_HEOs_TESTE.xlsx"


    df = pd.read_excel(file_name) 

    
    df[propriedade+"_minimo"] = None 
    df[propriedade+"_maximo"] = None 
    df[propriedade+"_soma"] = None 
    df[propriedade+"_media"] = None 
    df[propriedade+"_desvio"] = None 


    df.to_excel("Input_HEOs_TESTE.xlsx") 
    
    return
    
#criar_culunas_excel("electronegativity_ghosh")



def preencher_linhas (linha_inicial, linha_final) :

    file_name = "Input_HEOs_TESTE.xlsx"
    linha = linha_inicial
    
    propriedade_direta = "electronegativity_ghosh"
    
    
    df = pd.read_excel(file_name) 

    
    while linha < linha_final+1 :
     

        vetor = vetor_electronegativity_ghosh_atomos (linha)
      

        df.loc[linha, propriedade_direta+"_minimo"] = minimo(vetor)
        df.loc[linha, propriedade_direta+"_maximo"] = maximo(vetor)
        df.loc[linha, propriedade_direta+"_soma"] = soma(vetor)
        df.loc[linha, propriedade_direta+"_media"] = media(vetor)
        df.loc[linha, propriedade_direta+"_desvio"] = desvio(vetor)


        
            
        linha = linha +1
        
    df.to_excel("Input_HEOs_TESTE.xlsx") 
    
    return 


preencher_linhas(2,146)

