from mendeleev import element
import numpy as np 
import math
import itertools
from itertools import permutations
from itertools import combinations 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import math
from thermo_ml import parse
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np 
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re

#Atualizacao da funcao de scraping em 27/09/2023

def buscar_equacao(cation, dataframe):
    # Buscar a linha no DataFrame onde o cátion está localizado
    linha = dataframe[dataframe['Cation'] == cation]
    
    # Verificar se encontramos uma linha correspondente
    if len(linha) == 0:
        return "Cátion não encontrado no DataFrame!"
    
    # Retornar a equação da linha encontrada
    return linha.iloc[0]['Equacao']

def analisar_equacao(equacao):
    """
    This function receives a chemical equation string and returns the moles of cation, the moles of reacting oxygen,
    and their respective stoichiometric coefficients. It also returns the product and the moles of the product.
    
    :param equacao: str, The chemical equation string.
    :return: dict, A dictionary containing moles of cation, moles of oxygen, stoichiometric coefficients of cation
                  and oxygen, product, and moles of the product.
    """
    # Split the equation into reactants and products
    reagentes, produto = equacao.split(" -> ")
    
    # Find the coefficients and substances of the reactants
    reagentes = re.findall(r"(\d*)\s*([A-Za-z]+[0-9]*)", reagentes)
    
    # Initialize the variables
    mols_cation = mols_oxygen = coef_cation = coef_oxygen = 0
    coef_produto = 1  # Assume 1 if no coefficient is found in the product
    
    for coef, substancia in reagentes:
        coef = int(coef) if coef else 1  # If the coefficient is empty, assume 1
        if 'O2' == substancia:  # Exact match to avoid false positives
            mols_oxygen = coef
            coef_oxygen = 2  # Each mole of O2 has two oxygen atoms
        else:
            mols_cation = coef
            # Extract the subscript from the cation substance if present
            match = re.search(r"(\d+)$", substancia)
            coef_cation = int(match.group(1)) if match else 1  # If no subscript, assume 1
    
    # Find the coefficient of the product if present, and the substance of the product
    match = re.match(r"(\d*)\s*([A-Za-z0-9]+)", produto)
    if match:
        coef_str, produto = match.groups()
        coef_produto = int(coef_str) if coef_str else 1  # If the coefficient is empty, assume 1
    
    # The moles of the product are given by the coefficient of the product
    mols_produto = coef_produto
    
    return {
        "mols_cation": mols_cation,
        "mols_oxygen": mols_oxygen,
        "coef_cation": coef_cation,
        "coef_oxygen": coef_oxygen,
        "produto": produto,
        "mols_produto": mols_produto
    }

def aguardar_pagina_carregar(driver, timeout=20):
    WebDriverWait(driver, timeout).until(
        lambda x: driver.execute_script("return document.readyState === 'complete'")
    )

def funcao_scraping_atualizada(anion):

    navegador = webdriver.Chrome()
    navegador.get("https://www.crct.polymtl.ca/reacweb_plus.php")

    arquivo_excel = "Dataset_Fact_Sage_Updated_preenchido.xlsx"  
    df = pd.read_excel(arquivo_excel)  
    equacao = buscar_equacao(anion,df)
    
    resultado =  analisar_equacao(equacao)
    mols_cation = resultado["mols_cation"]
    mols_oxygen = resultado["mols_oxygen"]
    coef_cation = resultado["coef_cation"]
    coef_oxygen = resultado["coef_oxygen"]
    produto = resultado["produto"]
    mols_produto = resultado["mols_produto"]


    if int(mols_cation) != 1:
        nreact0 = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.NAME, "nreact[0]")))
        nreact0.send_keys(str(mols_cation))
        
    if int(coef_cation) == 1:
        react0 = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.NAME, "REACT[0]")))
        react0.send_keys(anion)
    else :
        react0 = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.NAME, "REACT[0]")))
        react0.send_keys(anion+str(coef_cation))  

    WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.NAME, "addReactant"))).click()
    
    
    if int(mols_oxygen) != 1:
        nreact1 = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.NAME, "nreact[1]")))
        nreact1.send_keys(str(mols_oxygen))
    if int(coef_oxygen) == 1:
        react1 = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.NAME, "REACT[1]")))
        react1.send_keys("O")
    else :
        react1 = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.NAME, "REACT[1]")))
        react1.send_keys("O"+str(coef_oxygen))    

    WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.NAME, "addProduct"))).click()
   

    if int(mols_produto) != 1:
        nprod = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.NAME, "nprod[0]")))
        nprod.send_keys(str(mols_produto))
        
    prod1 = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.NAME, "PROD[0]")))
    prod1.send_keys(produto)
    



        


    # Clicar no botão 'next' e esperar a página carregar
    WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.NAME, "next"))).click()
    aguardar_pagina_carregar(navegador)

    # Clicar novamente no botão 'next' e esperar a página carregar
    WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.NAME, "next"))).click()
    aguardar_pagina_carregar(navegador)

    # Preencher o campo 'Tinput' com "300"
    t_input = WebDriverWait(navegador, 10).until(EC.visibility_of_element_located((By.NAME, "Tinput")))
    t_input.clear()  # Limpar qualquer texto existente no campo.
    t_input.send_keys("300")

    # Clicar no botão 'calculateBTN' e esperar a página carregar
    WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.NAME, "calculateBTN"))).click()
    aguardar_pagina_carregar(navegador)

    # Obter todos os elementos com a classe 'result'
    lista_resultados = WebDriverWait(navegador, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "result")))

    entalpia_Oxido  = lista_resultados[3].text
    gibbs_Oxido  = lista_resultados[4].text
    entropia_Oxido  = lista_resultados[6].text
    delta_cp_Oxido  = lista_resultados[7].text 
 
    navegador.quit()    
    
    return entalpia_Oxido , entropia_Oxido , gibbs_Oxido , delta_cp_Oxido

    
def preencher_linhas (linha_inicial, linha_final) :

    file_name = "Dataset_Fact_Sage_Updated_preenchido.xlsx"
    linha = linha_inicial

    
    
    df = pd.read_excel(file_name) 

    
    while linha < linha_final+1 :
     
        
        
        anion =  str(df.at[linha, "Cation"])
        print(anion)
        
        try:
            entalpia_Oxido , entropia_Oxido , gibbs_Oxido , delta_cp_Oxido = funcao_scraping_atualizada (anion)
        except:
            linha = linha -1
            
        finally:
        
            try:
                df.loc[linha,"Entropia"] = entropia_Oxido
                df.loc[linha,"Entalpia"] = entalpia_Oxido
                df.loc[linha,"Gibbs"] = gibbs_Oxido
                df.loc[linha,"delta_cp"] = delta_cp_Oxido
            except:
                pass

            if str(anion).find("W") != -1:
                break
            df.to_excel("Dataset_Fact_Sage_Updated_preenchido.xlsx")     
            linha = linha +1
        
    
    
    return 
    
   
​

​


preencher_linhas(0,49)


