from mendeleev import element
import numpy as np 
import math
import itertools
from itertools import permutations
from itertools import combinations 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import math
from thermo_ml import parse
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np 

def minimo (vetor) :
    return min(vetor)

def maximo (vetor) :
    return max(vetor)

def soma (vetor) :
    return sum(vetor)

def media (vetor) :
    return sum(vetor)/len(vetor)

def desvio (vetor) :
    return np.std(vetor)

# Load the dataset
filename = 'Dataset_Fact_Sage_Updated_preenchido.xlsx'
df = pd.read_excel(filename)

def buscar_valor_cation(cation, propriedade):

    print(cation)
    row = df.loc[df['Cation'] == cation]
    
    if row.empty:
        return f"Cátion {cation} não encontrado no dataset."
    
    if propriedade not in df.columns:
        return f"Propriedade {propriedade} não encontrada no dataset."

    return float (row[propriedade].values[0])


def vetor_valor (linha, propriedade):

    wb = load_workbook('Atomic-features_.xlsx')
    ws = wb['Sheet1']
        
    atomos = ws["B"+ str(linha)].value
        

    atomos = atomos.replace (" ","")
    atomos = atomos.replace ("'","")
    atomos = atomos.replace ("[","")
    atomos = atomos.replace ("]","")
    atomos = atomos.split(",")
    
    if 'O' in atomos:
        atomos = [cation for cation in atomos if cation != 'O']
    
    print(atomos)
    vetor_propriedade_da_linha = []

    for elemento in atomos:
       
         vetor_propriedade_da_linha.append(buscar_valor_cation(elemento, propriedade))

    return vetor_propriedade_da_linha
    





def preencher_linhas (linha_inicial, linha_final) :

    file_name = "Atomic-features_.xlsx"
    linha = linha_inicial
    

    
    df = pd.read_excel(file_name) 

    
    while linha < linha_final+1 :
     
        lista_proprieades = ['entalpia-oxidos','gibbs-oxidos','entropia-oxidos','deltaCp-oxidos']
        
    
        for priprieade in lista_proprieades:

            vetor = vetor_valor(linha, priprieade)

            df.loc[linha, priprieade+"_minimo"] = minimo(vetor)
            df.loc[linha, priprieade+"_maximo"] = maximo(vetor)
            df.loc[linha, priprieade+"_soma"] = soma(vetor)
            df.loc[linha, priprieade+"_media"] = media(vetor)
            df.loc[linha, priprieade+"_desvio"] = desvio(vetor)


        
            
        linha = linha +1
        print(linha)
        
    df.to_excel("Atomic-features_termo.xlsx") 
    
    return 




preencher_linhas(2,170)

