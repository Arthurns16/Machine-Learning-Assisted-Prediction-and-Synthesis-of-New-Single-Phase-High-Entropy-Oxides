from mendeleev import element
import numpy as np 
import math
import itertools
from itertools import permutations
from itertools import combinations 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import math
from thermo_ml import parse
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import pandas as pd

import numpy as np 

def minimo (vetor) :
    return min(vetor)

def maximo (vetor) :
    return max(vetor)

def soma (vetor) :
    return sum(vetor)

def media (vetor) :
    return sum(vetor)/len(vetor)

def desvio (vetor) :
    return np.std(vetor)



def roman_to_int(romano):

    romano = romano.replace("S", "")
    romano = romano.replace("Q", "")
    romano = romano.replace("P", "")
    romano = romano.replace("Y", "")
    
    nums = {'M':1000,
            'D':500,
            'C':100,
            'L':50,
            'X':10,
            'V':5,
            'I':1}
    sum = 0
    for i in range(len(romano)):
       
            value = nums[romano[i]]
            if i+1 < len(romano) and nums[romano[i+1]] > value:
                sum -= value
            else: sum += value
        
        
    return sum
    

def coordenacao_medio_elemento (atomo):

    atomo = atomo
    atomo = element(atomo)    
    
    soma = 0 
    numero_coordenacoes_possiveis = 0
    
    for coordenacao in atomo.ionic_radii :
        numero_coordenacoes_possiveis = numero_coordenacoes_possiveis +1
        soma = soma +roman_to_int(coordenacao.coordination )
   
    return float(soma)/float(numero_coordenacoes_possiveis)
    



    

def vetor_coordenacao_atomos (linha) :

   
    wb = load_workbook('Input_HEOs_TESTE.xlsx')
    ws = wb['Sheet1']
        
    atomos = ws["B"+ str(linha)].value
        

    atomos = atomos.replace (" ","")
    atomos = atomos.replace ("'","")
    atomos = atomos.replace ("[","")
    atomos = atomos.replace ("]","")
    atomos = atomos.split(",")

    vetor_propriedade_da_linha = []

    for elemento in atomos:
        try: 
            vetor_propriedade_da_linha.append (float(coordenacao_medio_elemento (elemento)))
        except:
            valor = 0
        
    return vetor_propriedade_da_linha




def criar_culunas_excel (propriedade) : 

    file_name = "Input_HEOs_TESTE.xlsx"


    df = pd.read_excel(file_name) 

    
    df[propriedade+"_minimo"] = None 
    df[propriedade+"_maximo"] = None 
    df[propriedade+"_soma"] = None 
    df[propriedade+"_media"] = None 
    df[propriedade+"_desvio"] = None 


    df.to_excel("Input_HEOs_TESTE.xlsx") 




def preencher_linhas (linha_inicial, linha_final) :

    file_name = "Input_HEOs_TESTE.xlsx"
    linha = linha_inicial
    
    propriedade_direta = "coordenacao"
    
    
    df = pd.read_excel(file_name) 

    
    while linha < linha_final+1 :
     

        df.loc[linha, propriedade_direta+"_minimo"] = minimo(vetor_coordenacao_atomos (linha))
        df.loc[linha, propriedade_direta+"_maximo"] = maximo(vetor_coordenacao_atomos (linha))
        df.loc[linha, propriedade_direta+"_soma"] = soma(vetor_coordenacao_atomos (linha))
        df.loc[linha, propriedade_direta+"_media"] = media(vetor_coordenacao_atomos (linha))
        df.loc[linha, propriedade_direta+"_desvio"] = desvio(vetor_coordenacao_atomos (linha))

        
            
        linha = linha +1
        
    df.to_excel("Input_HEOs_TESTE.xlsx") 
    
    return 


preencher_linhas(2,146)







